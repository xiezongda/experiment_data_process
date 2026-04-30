# 写入 Excel
# 每个 CSV 对应一个 sheet
# 自动处理 sheet 名非法字符
# 自动处理 sheet 重名
# 可以创建新 Excel，也可以覆盖已有 sheet


from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook              #用于读取和修改 Excel 文件
from openpyxl.styles import Alignment, Font     #导入字体和对齐样式
from openpyxl.utils import get_column_letter    #用于将列索引转换为 Excel 列字母


def safe_sheet_name(name: str) -> str:
    """
    Excel sheet 名称限制：
    - 最大 31 个字符
    - 不能包含 []:*?/\\
    """
    name = str(name).strip()
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)  #r表示原始字符串，避免转义

    if name == "":
        name = "sheet"

    return name[:31]


def make_unique_sheet_name(
    desired_name: str,
    existing_names: set[str],
) -> str:
    """
    如果 sheet 名已经存在，就自动添加后缀，从 _2 开始。
    """
    desired_name = safe_sheet_name(desired_name)

    if desired_name not in existing_names:
        return desired_name

    base = desired_name[:25]

    index = 2
    while True:
        candidate = safe_sheet_name(f"{base}_{index}")
        if candidate not in existing_names:
            return candidate
        index += 1


def get_existing_sheet_names(excel_path: Path) -> set[str]:
    """
    获取已有 Excel 文件中的 sheet 名。
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        return set()

    workbook = load_workbook(excel_path, read_only=True)   #只读模式打开，效率更高
    return set(workbook.sheetnames)                        #返回一个 set，方便快速查重


def autosize_excel_columns(excel_path: Path) -> None:
    """
    简单自动调整 Excel 列宽。
    """
    workbook = load_workbook(excel_path)

    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)# 获取列字母，例如 'A', 'B', 'C'...

            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    workbook.save(excel_path)


def write_tables_to_excel(
    excel_path: Path,
    tables: dict[str, pd.DataFrame],
    mode: str = "replace_file",
    auto_width: bool = True,
) -> None:
    """
    将多个 DataFrame 写入一个 Excel 文件。

    参数
    ----
    excel_path:
        输出 Excel 路径。

    tables:
        形如：
        {
            "acf_horizontal": df1,
            "acf_vertical": df2,
            "psd_horizontal": df3,
        }

    mode:
        - "replace_file":
            如果 Excel 已存在，整个文件重新生成。
            适合当前阶段，最简单、最稳定。

        - "append_new_sheets":
            如果 Excel 已存在，保留原 sheet，新数据写入新 sheet。
            如果 sheet 重名，会自动加 _2、_3。

    auto_width:
        是否自动调整列宽。
    """
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if mode not in {"replace_file", "append_new_sheets"}:
        raise ValueError(
            "mode 只能是 'replace_file' 或 'append_new_sheets'"
        )

    if mode == "replace_file":
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in tables.items():
                df.to_excel(
                    writer,
                    sheet_name=safe_sheet_name(sheet_name),
                    index=False,
                )

    elif mode == "append_new_sheets":
        existing_names = get_existing_sheet_names(excel_path)

        if excel_path.exists():
            writer_mode = "a"
        else:
            writer_mode = "w"

        with pd.ExcelWriter(
            excel_path,
            engine="openpyxl",
            mode=writer_mode,
            if_sheet_exists="new" if writer_mode == "a" else None,
        ) as writer:
            for sheet_name, df in tables.items():
                final_sheet_name = make_unique_sheet_name(
                    desired_name=sheet_name,
                    existing_names=existing_names,
                )
                existing_names.add(final_sheet_name)

                df.to_excel(
                    writer,
                    sheet_name=final_sheet_name,
                    index=False,
                )

    if auto_width:
        autosize_excel_columns(excel_path)